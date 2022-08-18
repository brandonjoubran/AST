#!/bin/sh

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import requests
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from time import sleep
import os
from datetime import datetime
#from Formulas import Formulas
import tkinter as tk
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup
import gspread
from oauth2client.service_account import ServiceAccountCredentials

class Formulas:

    @staticmethod
    def gap_up_perc_open_formula(open_price, prev_close):
        try:
            open_price_int = float(open_price)
            prev_close_int = float(prev_close)
            if(prev_close_int == 0):
                return 0
            return ((open_price_int - prev_close_int) / prev_close_int) * 100
        except:
            print("Error occured in gap_up_perc_open_formula(). Values given were open_price={} and prev_close={}".format(open_price, prev_close))
        return 'N/A'


    @staticmethod
    def gap_up_perc_premarket_formula(prev_close, premarket_high):
        #premarket_high = self.split_premarket_high(premarket_high)
        try:
            premarket_high_float = float(premarket_high)
            prev_close_float = float(prev_close)
            if (prev_close_float == 0):
                return 0
            return ((premarket_high_float - prev_close_float)/prev_close_float)*100
        except:
            print("Error occured in gap_up_perc_premarket_formula(). Values given were prev_close={} and premarket_high={}".format(prev_close, premarket_high))
        return 'N/A'

    @staticmethod
    def gap_perc_maintained_by_open(gap_up_perc_open, gap_up_perc_premarket):
        try:
            if(gap_up_perc_premarket == 0):
                return 0
            result = (gap_up_perc_open/gap_up_perc_premarket)*100
            return result
        except:
            print("Error occured in gap_perc_maintained_by_open(). Values given were gap_up_perc_open={} and gap_up_perc_premarket={}".format(
                gap_up_perc_open, gap_up_perc_premarket
            ))
        return 'N/A'

    @staticmethod
    def spike_perc(days_high, open_price):
        try:
            open_price_float = float(open_price)
            days_high_float = float(days_high)
            if(open_price_float == 0):
                return 0
            return ((days_high_float - open_price_float) / open_price_float) * 100
        except:
            print("Error occured in spike_perc(). Values given were days_high={} and open_price={}".format(days_high, open_price))
        return 'N/A'

    @staticmethod
    def fail_perc(days_low, open_price):
        try:
            open_price_float = float(open_price)
            days_low_float = float(days_low)
            if(open_price_float == 0):
                return 0
            return ((open_price_float - days_low_float) / open_price_float) * 100
        except:
            print("Error occured in fail_perc(). Values given were days_low={} amd open_price={}".format(days_low, open_price))
        return 'N/A'

    @staticmethod
    def perc_of_float_trade(float, vol):
        try:
            if(float == 0):
                return 0
            return (vol/float)*100
        except:
            print("Error occured in perc_of_float_trade(). Values given were float={} amd vol={}".format(float, vol))
        return 'N/A'

    @staticmethod
    def pullback_from_pm_high_to_open(premarket_high, open_price):
        try:
            open_price_float = float(open_price)
            premarket_high_float = float(premarket_high)
            if(premarket_high_float == 0):
                return 0
            return ((premarket_high_float - open_price_float)/premarket_high_float)*100
        except:
            print("Error occured in pullback_from_pm_high_to_open(). Values given were premarket_high={} amd open_price={}".format(premarket_high, open_price))
        return 'N/A'

class StockScraper:
    '''NOTE: To run the program, at the bottom you just have to create an instance of the class, and everything should get done.
    
    Initializing some data and starting the program
    '''
   
    def __init__(self, stocks, author, uid):
        #IMPORTANT: Switch path_to_cd3 to your own path
        self.path_to_cd3 = '/Users/brandonjoubran/Downloads/chromedriver 6'
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.stocks = stocks
        self.filename = author + ".xlsx"
        self.path = ""
        # Desired path to save excel sheet. If nothing is entered, the file will be saved here. Else, the path will be saved for future use
        self.custom_path = ""
        # Dictionary for individual stock
        self.dict = {}

        #Dictionary for all stocks (this will be used for excel)
        self.final_dict = {}

        self.barchat_api = 'cb6217f02158e9fab503490d6c993c45'
        self.date = datetime.today().strftime('%Y-%m-%d')

        self.error_msg = ""

        # Used for Google Spreadsheets 

        scope = ["https://spreadsheets.google.com/feeds", 'https://www.googleapis.com/auth/spreadsheets',
                 "https://www.googleapis.com/auth/drive.file", "https://www.googleapis.com/auth/drive"]
        creds = ServiceAccountCredentials.from_json_keyfile_name('ast.json', scope)
        client_sp = gspread.authorize(creds)
        self.database_sheet = client_sp.open('Clients').worksheet(uid)

        #These are the categories in column A. If you want to rearrange the rows, change the value assosciated with it. i.e. to
        #move Stock to row 25, change the value 1 to 25 and make sure all values in the dictionary are unique. If you want to change the
        #names, change it here and at the bottom of search_yahoo/nasdaq. i.e. to change 'Stock' to 'Name', change it 'Stock':1 to 'Name':1
        #here, and at the bottom of search_yahoo, change self.dict['Stock'] = keys to self.dict['Name'] = keys

        self.categories = {'Stock': 1, 'Closing Price': 8, 'Days High': 3, 'Days Low': 5, 'Previous Close': 9, 'Open Price': 7, 'Shares Float': 10, 'Short Float %': 11, 'Inst. Own %': 12, 'Pre-Market High': 14, 'Time of pre-market high': 15, 'Pre-Market Low': 16, 'Time of pre-market low': 17, 'Market Cap': 18, 'Spike % from open to HOD': 19, 'Gap up % (to open)': 20, 'Gap up % (to Pre-Market High)': 21, 'Gap % maintained by open': 22, 'Fail %': 24, 'Volume at 1m': 32, 'Volume at 2m': 35, 'Volume at 5m': 41, 'Volume at 15m': 47, 'Volume at 30m': 56, 'Pre-Market Volume': 13, 'Volume at 1m (Inclusive of PM vol)': 33, 'Volume at 2m (Inclusive of PM vol)': 36, 'Volume at 5m (Inclusive of PM vol)': 42, 'Volume at 15m (Inclusive of PM vol)': 48, 'Volume at 30m (Inclusive of PM vol)': 57, '% of Float Trades (End of Pre-Market)': 25, '% of Float Trades (After 1m)': 34, '% of Float Trades (After 2m)': 37, 'Volume at 3m': 38, 'Volume at 3m (Inclusive of PM vol)': 39, '% of Float Trades (After 3m)': 40, 'Volume at 10m': 44, 'Volume at 10m (Inclusive of PM vol)': 45, '% of Float Trades (After 10m)': 46, 'Volume at 20m': 50, 'Volume at 20m (Inclusive of PM vol)': 51, '% of Float Trades (After 20m)': 52, 'Volume at 25m': 53, 'Volume at 25m (Inclusive of PM vol)': 54, '% of Float Trades (After 25m)': 55, '% of Float Trades (After 5m)': 43, '% of Float Trades (After 15m)': 49, '% of Float Trades (After 30m)': 58, 'Pullback from PM high to open': 23, 'Date': 2, 'Time of HOD': 4, 'Time of LOD': 6, 'End of Day Volume': 31, '52 Week High': 26, '52 Week Low': 27, '50-Day Moving Average': 28, '200-Day Moving Average': 29, 'Avg Vol (3 month)': 30}
        self.start()

    def start(self, key=''):

        '''
        Initializes WebDriver and loops through requested stocks. Once done, move info to Excel spreadsheet
        '''

        for cat in list(self.categories.keys()):
            self.dict[cat] = 'N/A'

        options = Options()
        options.headless = True
        user_agent = 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.50 Safari/537.36'
        options.add_argument('user-agent={0}'.format(user_agent))
        driver = webdriver.Chrome(self.path_to_cd3, chrome_options=options)


        #This just checks if the values self.categories is unique. If not it will print which values are repeated.
        if(len(list(set(self.categories.values()))) != len(list(range(1, len(self.categories.values())+1)))):
            print("Make sure the values in self.categories are unique. Fix and run again.")
            print("Values are: {}".format(list(set(self.categories.values()))))
            print("Should be:  {}".format(list(self.categories.values())))
            dup = []
            for i in self.categories.values():
                if(i in dup):
                    print("More than one item at row {}".format(i))
                dup += [i]
            return

        #to_remove is a list of stocks that error when the program is running. They are removed after to avoid error
        to_remove = []

        if(self.stocks == [] and key == ''):
            print("The file is empty and no key was given.")
        elif(self.stocks != []):
            for stock in self.stocks:
                print('Now doing: {}'.format(stock))
                if(stock == '' or stock == ' ' or stock == None or stock[0] == ' '):
                    print("Was empty, none, or started with a space: {}".format(stock))
                    to_remove += [stock]
                    continue

                try:
                    self.calculations(stock, driver)
                except:
                    print('An error occured when checking {}. Double check stock names.'.format(stock))
                    #self.popup('An error occured when checking {}. Double check stock names.'.format(stock))
                    if(self.dict != {}):
                        to_remove += [stock]
                        self.dict = {}

            for stock in to_remove:
                self.stocks.remove(stock)
        else:
            self.calculations(key, driver)

        driver.quit()
        print("Now transferring to excel")
        self.setup_excel(self.filename)
        self.excel(self.final_dict)

    def popup(self, msg):
        popup = tk.Tk()
        popup.title('Stock Buddy')
        label = tk.Label(popup, text=msg)
        label.pack()

    '''def to_directory(self):
        from_path = self.path_textbox.get('1.0', tk.END).strip('\n').replace(" ", "")
        if (from_path != ""):
            try:
                os.chdir(from_path)
            except:
                pass
                #self.popup("Error with path.")
        print(from_path)'''

    def calculations(self, key, driver):

        '''
        Starting search for Yahoo and Nasdaq and using Formulas class for calculations
        '''

        for cat in list(self.categories.keys()):
            self.dict[cat] = 'N/A'

        #key is the stock code i.e. AAPL
        self.search_yahoo(key, driver)
        print("Finished Searching Yahoo")
        self.search_nasdaq(key, driver)

        # Doing calculations 
        gap_up_perc_open = Formulas.gap_up_perc_open_formula(self.dict['Open Price'], self.split_comma(self.dict['Previous Close']))
        gap_up_perc_premarket = Formulas.gap_up_perc_premarket_formula(self.split_comma(self.dict['Previous Close']), self.split_comma(self.dict['Pre-Market High']))
        self.dict['Gap up % (to open)'] = gap_up_perc_open
        self.dict['Gap up % (to Pre-Market High)'] = gap_up_perc_premarket
        self.dict['Gap % maintained by open'] = Formulas.gap_perc_maintained_by_open(gap_up_perc_open, gap_up_perc_premarket)
        self.dict['Spike % from open to HOD'] = Formulas.spike_perc(self.split_comma(self.dict['Days High']), self.dict['Open Price'])
        self.dict['Fail %'] = Formulas.fail_perc(self.split_comma(self.dict['Days Low']), self.dict['Open Price'])
        self.dict['Pullback from PM high to open'] = Formulas.pullback_from_pm_high_to_open(self.split_comma(self.dict['Pre-Market High']), self.dict['Open Price'])
        self.dict['Date'] = self.date
        row = [self.dict['Stock'], self.dict['Date'], self.dict['Closing Price'], self.dict['Market Cap'], self.dict['Spike % from open to HOD'], self.dict['Fail %'], self.dict['Shares Float'], self.dict['Short Float %'], self.dict['Inst. Own %']]
        self.update_database(row)
        self.final_dict[key] = self.dict
        self.dict = {}


    def update_database(self, row):
        self.database_sheet.append_row(row)

    def search_yahoo(self, keys, driver):

        '''
        Scraping Yahoo for information
        '''
        url = 'https://ca.finance.yahoo.com/quote/{}'.format(keys)

        result = requests.get(url)
        src = result.content
        soup = BeautifulSoup(src, 'lxml')
        try:
            #price_text = soup.select("div span[data-reactid*='14']")[0].text
            price_text = "N/A"
            previous_close_text = soup.find_all("td", class_="Ta(end) Fw(600) Lh(14px)")[0].text
            open_price_text = soup.find_all("td", class_="Ta(end) Fw(600) Lh(14px)")[1].text
            volume_text = soup.find_all("td", class_="Ta(end) Fw(600) Lh(14px)")[6].text
            market_cap_text = soup.find_all("td", class_="Ta(end) Fw(600) Lh(14px)")[8].text
            days_range_text = soup.find_all("td", class_="Ta(end) Fw(600) Lh(14px)")[4].text
            self.split_day_range(days_range_text)
            url_statistics = 'https://ca.finance.yahoo.com/quote/{}/key-statistics?p={}'.format(keys, keys)
            result = requests.get(url_statistics)
            src = result.content
            soup = BeautifulSoup(src, 'html.parser')
            self.dict['Stock'] = keys
            self.dict['Closing Price'] = price_text
            self.dict['Previous Close'] = previous_close_text
            self.dict['Open Price'] = open_price_text
            self.dict['Market Cap'] = market_cap_text
            self.dict['End of Day Volume'] = volume_text
        except:
            pass

        try:
            held_by_inst_text = soup.find_all("td", class_="Fw(500) Ta(end) Pstart(10px) Miw(60px)")[21].text
            shares_float_text = soup.find_all("td", class_="Fw(500) Ta(end) Pstart(10px) Miw(60px)")[19].text
            short_perc_float_text = soup.find_all("td", class_="Fw(500) Ta(end) Pstart(10px) Miw(60px)")[24].text
            fifty_two_week_high_text = soup.find_all("td", class_="Fw(500) Ta(end) Pstart(10px) Miw(60px)")[12].text
            fifty_two_week_low_text = soup.find_all("td", class_="Fw(500) Ta(end) Pstart(10px) Miw(60px)")[13].text
            fifty_day_moving_avg = soup.find_all("td", class_="Fw(500) Ta(end) Pstart(10px) Miw(60px)")[14].text
            two_hundred_day_moving_avg = soup.find_all("td", class_="Fw(500) Ta(end) Pstart(10px) Miw(60px)")[15].text
            avg_vol_3_months = soup.find_all("td", class_="Fw(500) Ta(end) Pstart(10px) Miw(60px)")[16].text
            self.dict['Inst. Own %'] = held_by_inst_text
            self.dict['Shares Float'] = shares_float_text
            self.dict['Short Float %'] = short_perc_float_text
            self.dict['52 Week High'] = fifty_two_week_high_text
            self.dict['52 Week Low'] = fifty_two_week_low_text
            self.dict['50-Day Moving Average'] = fifty_day_moving_avg
            self.dict['200-Day Moving Average'] = two_hundred_day_moving_avg
            self.dict['Avg Vol (3 month)'] = avg_vol_3_months
        except:
            pass

    def search_nasdaq(self, keys, driver):
        '''
        Scraping NASDAQ site for pre-market info
        '''
        url = 'https://www.nasdaq.com/market-activity/stocks/{}/pre-market'.format(keys)
        wait = 15
        print(url)
        driver.get(url)
        print('past url')
        sleep(10)

        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')

        try:
            pre_market_volume_text = soup.find_all('td', class_='pre-market-quote-info__cell')[0].text
            self.dict['Pre-Market Volume'] = self.split_comma(pre_market_volume_text)
        except:
            pass
        try:
            pre_market_high_text = soup.find_all('td', class_='pre-market-quote-info__cell')[1].text
            self.dict['Pre-Market High'] = self.split_premarket_high(pre_market_high_text)
        except:
            pass

        try:
            pre_market_low_text = soup.find_all('td', class_='pre-market-quote-info__cell')[2].text
            self.dict['Pre-Market Low'] = self.split_premarket_low(pre_market_low_text)
        except:
            pass

        tens = {'k': 10e2, 'K': 10e2, 'm': 10e5, 'M': 10e5, 'b':10e8, 'B': 10e8}
        f = lambda x: int(float(x[:-1]) * tens[x[-1]])

        if(self.dict['Pre-Market Volume'] == 'N/A' or self.dict['Shares Float'] == 'N/A'):
            print("Can't compute % of Float Trades due to no value of pre-market volume")
            self.dict['% of Float Trades (End of Pre-Market)'] = 'N/A'
            return
        self.dict['% of Float Trades (End of Pre-Market)'] = (float(self.dict['Pre-Market Volume'])/f(self.dict['Shares Float']))*100

    def volumes_at(self, key):

        '''
        WARNING: This function is depracated due to the API it utilized becoming unavailable

        Utilized an API to retrieve data about a stock within x minutes of the market opening.
        Get the volumes at 1min, 2min, ..., 30min.
        '''

        key1 = 'cb6217f02158e9fab503490d6c993c45'
        key2 = '664c45e733b1f7c2117e117480771ca6'
        try:
            url = 'https://marketdata.websol.barchart.com/getHistory.json?apikey={}&symbol={}&type=minutes&startDate={}'.format(key1, key, self.date)
            result = requests.get(
                url
            )

            url2 = 'https://marketdata.websol.barchart.com/getHistory.json?apikey={}&symbol={}&type=daily&startDate={}'.format(
                key1, key, self.date)
            result2 = requests.get(
                url2
            )

            try:
                high = result2.json()['results'][0]['high']
            except:
                high = 'N/A'
            try:
                low = result2.json()['results'][0]['low']
            except:
                low = 'N/A'
        except:
            url = 'https://marketdata.websol.barchart.com/getHistory.json?apikey={}&symbol={}&type=minutes&startDate={}'.format(
                key2, key, self.date)
            result = requests.get(
                url
            )

            url2 = 'https://marketdata.websol.barchart.com/getHistory.json?apikey={}&symbol={}&type=daily&startDate={}'.format(
                key2, key, self.date)
            result2 = requests.get(
                url2
            )

            try:
                high = result2.json()['results'][0]['high']
            except:
                high = 'N/A'
            try:
                low = result2.json()['results'][0]['low']
            except:
                low = 'N/A'

        try:
            self.dict['Open Price'] = result2.json()['results'][0]['open']
            self.dict['Closing Price'] = result2.json()['results'][0]['close']
        except:
            self.dict['Open Price'] = 'N/A'
        print(url)
        if(result.json()['results'] == None):
            print("Can't compute volumes at x minute due to lack of data. On weekends no data is provided")
            for i in range(31):
                if(i == 1 or i == 2 or i == 3 or i == 5 or i == 10 or i == 15 or i == 20 or i == 25 or i == 30):
                    print()
                    self.dict['Volume at {}m'.format(i)] =  'N/A'
                    self.dict['Volume at {}m (Inclusive of PM vol)'.format(i)] = 'N/A'
                    self.dict['% of Float Trades (After {}m)'.format(i)] = 'N/A'
            return

        x = 0
        tens = {'k': 10e2, 'K': 10e2, 'm': 10e5, 'M': 10e5, 'b':10e8, 'B': 10e8}
        f = lambda x: int(float(x[:-1]) * tens[x[-1]])

        try:
            shares_float = f(self.dict['Shares Float'])
        except:
            shares_float = self.dict['Shares Float']
        min = 30
        error_msg = ""

        min_1 = self.getSecond('9:30:00')
        min_2 = self.getSecond('9:31:00')
        min_3 = self.getSecond('9:32:00')
        min_5 = self.getSecond('9:34:00')
        min_10 = self.getSecond('9:39:00')
        min_15 = self.getSecond('9:44:00')
        min_20 = self.getSecond('9:49:00')
        min_25 = self.getSecond('9:54:00')
        min_30 = self.getSecond('09:59:00')

        #Finds the time of HOD/LOD by looking at the trades/minute, and finding the closest ones to the HOD and LOD and using that time. This is to avoid errors, as the low can be, for example, 4.52 but at the time of LOD it trades at 4.48 or 4.1555
        closest_high = float('-inf')
        closest_high_time = ''
        closest_low = float('inf')
        closest_low_time = ''
        x = 0
        for value in result.json()['results']:
            x += int(value['volume'])
            time = self.getTime(value['timestamp'])

            if self.getSecond(time) <= min_1:
                self.dict['Volume at 1m'] = x
                try:
                    self.dict['Volume at 1m (Inclusive of PM vol)'] = x + int(self.dict['Pre-Market Volume'])
                except:
                    self.dict['Volume at 1m (Inclusive of PM vol)'] = x + 0
                try:
                    self.dict['% of Float Trades (After 1m)'] = Formulas.perc_of_float_trade(shares_float, x + int(self.dict['Pre-Market Volume']))
                except:
                    self.dict['% of Float Trades (After 1m)'] = Formulas.perc_of_float_trade(shares_float, x + 0)

            if self.getSecond(time) <= min_2:
                self.dict['Volume at 2m'] = x
                try:
                    self.dict['Volume at 2m (Inclusive of PM vol)'] = x + int(self.dict['Pre-Market Volume'])
                except:
                    self.dict['Volume at 2m (Inclusive of PM vol)'] = x + 0
                try:
                    self.dict['% of Float Trades (After 2m)'] = Formulas.perc_of_float_trade(shares_float, x + int(self.dict['Pre-Market Volume']))
                except:
                    self.dict['% of Float Trades (After 2m)'] = Formulas.perc_of_float_trade(shares_float, x + 0)

            if self.getSecond(time) <= min_3:
                self.dict['Volume at 3m'] = x
                try:
                    self.dict['Volume at 3m (Inclusive of PM vol)'] = x + int(self.dict['Pre-Market Volume'])
                except:
                    self.dict['Volume at 3m (Inclusive of PM vol)'] = x + 0
                try:
                    self.dict['% of Float Trades (After 3m)'] = Formulas.perc_of_float_trade(shares_float, x + int(self.dict['Pre-Market Volume']))
                except:
                    self.dict['% of Float Trades (After 3m)'] = Formulas.perc_of_float_trade(shares_float, x + 0)

            if self.getSecond(time) <= min_5:
                self.dict['Volume at 5m'] = x
                try:
                    self.dict['Volume at 5m (Inclusive of PM vol)'] = x + int(self.dict['Pre-Market Volume'])
                except:
                    self.dict['Volume at 5m (Inclusive of PM vol)'] = x + 0
                try:
                    self.dict['% of Float Trades (After 5m)'] = Formulas.perc_of_float_trade(shares_float, x + int(self.dict['Pre-Market Volume']))
                except:
                    self.dict['% of Float Trades (After 5m)'] = Formulas.perc_of_float_trade(shares_float, x + 0)

            if self.getSecond(time) <= min_10:
                self.dict['Volume at 10m'] = x
                try:
                    self.dict['Volume at 10m (Inclusive of PM vol)'] = x + int(self.dict['Pre-Market Volume'])
                except:
                    self.dict['Volume at 10m (Inclusive of PM vol)'] = x + 0
                try:
                    self.dict['% of Float Trades (After 10m)'] = Formulas.perc_of_float_trade(shares_float, x + int(self.dict['Pre-Market Volume']))
                except:
                    self.dict['% of Float Trades (After 10m)'] = Formulas.perc_of_float_trade(shares_float, x + 0)

            if self.getSecond(time) <= min_15:
                self.dict['Volume at 15m'] = x
                try:
                    self.dict['Volume at 15m (Inclusive of PM vol)'] = x + int(self.dict['Pre-Market Volume'])
                except:
                    self.dict['Volume at 15m (Inclusive of PM vol)'] = x + 0
                try:
                    self.dict['% of Float Trades (After 15m)'] = Formulas.perc_of_float_trade(shares_float, x + int(self.dict['Pre-Market Volume']))
                except:
                    self.dict['% of Float Trades (After 15m)'] = Formulas.perc_of_float_trade(shares_float, x + 0)
            
            if self.getSecond(time) <= min_20:
                self.dict['Volume at 20m'] = x
                try:
                    self.dict['Volume at 20m (Inclusive of PM vol)'] = x + int(self.dict['Pre-Market Volume'])
                except:
                    self.dict['Volume at 20m (Inclusive of PM vol)'] = x + 0
                try:
                    self.dict['% of Float Trades (After 20m)'] = Formulas.perc_of_float_trade(shares_float, x + int(self.dict['Pre-Market Volume']))
                except:
                    self.dict['% of Float Trades (After 20m)'] = Formulas.perc_of_float_trade(shares_float, x + 0)
            
            if self.getSecond(time) <= min_25:
                self.dict['Volume at 25m'] = x
                try:
                    self.dict['Volume at 25m (Inclusive of PM vol)'] = x + int(self.dict['Pre-Market Volume'])
                except:
                    self.dict['Volume at 25m (Inclusive of PM vol)'] = x + 0
                try:
                    self.dict['% of Float Trades (After 25m)'] = Formulas.perc_of_float_trade(shares_float, x + int(self.dict['Pre-Market Volume']))
                except:
                    self.dict['% of Float Trades (After 25m)'] = Formulas.perc_of_float_trade(shares_float, x + 0)
            
            if self.getSecond(time) <= min_30:
                self.dict['Volume at 30m'] = x
                try:
                    self.dict['Volume at 30m (Inclusive of PM vol)'] = x + int(self.dict['Pre-Market Volume'])
                except:
                    self.dict['Volume at 30m (Inclusive of PM vol)'] = x + 0
                try:
                    self.dict['% of Float Trades (After 30m)'] = Formulas.perc_of_float_trade(shares_float, x + int(self.dict['Pre-Market Volume']))
                except:
                    self.dict['% of Float Trades (After 30m)'] = Formulas.perc_of_float_trade(shares_float, x + 0)

            if(abs(low - value['low']) <= abs(low - closest_low) and value['low'] <= closest_low):
                closest_low = value['low']
                closest_low_time = value['timestamp']
                self.dict['Time of LOD'] = self.getTime(value['timestamp'])
            
            if (abs(high - value['high']) <= abs(high - closest_high) and value['high'] >= closest_high):
                closest_high = value['high']
                closest_high_time = value['timestamp']
                self.dict['Time of HOD'] = self.getTime(value['timestamp'])

        if(error_msg != ""):
            pass


    def getTime(self, time):
        time = time.split('T')[1]
        time = time.split('-')[0]
        return time

    def getSecond(self, time):
        timesplit = time.split(":")
        hour = float(timesplit[0])
        minute = float(timesplit[1])
        seconds = float(timesplit[2])
        return (hour * 3600) + (minute * 60) + seconds

    def split_premarket_high(self, str):
        str = str[1:]
        str = str.split('(')
        self.dict['Pre-Market High'] = str[0]
        self.dict['Time of pre-market high'] = str[1][:-1]
        return str[0]

    def split_day_range(self, str):
        str = str.split('-')
        self.dict['Days High'] = str[1]
        self.dict['Days Low'] = str[0]
        return str

    def split_comma(self, str):
        str = str.replace(',', '')
        return str

    def split_premarket_low(self, str):
        str = str[1:]
        str = str.split('(')
        self.dict['Time of pre-market low'] = str[1][:-1]
        return str[0]

    def from_file(self, file):

        #Read the stocks from the file

        try:
            f = open(file, "r")
        except:
            print("Make sure file is in the same folder as the code!")

        for stock in f:
            self.stocks += [stock.rstrip()]

        print("Stocks retrieved from {}: {}".format(file, self.stocks))


    def import_workbook(self, file=''):
        #Try opening existing excel sheet (not currently being used)

        try:
            wb = load_workbook(file)
            return wb
        except:
            self.setup_excel()

    def setup_excel(self, excel_file_name):
        #Initialize excel sheet, these are the categories

        wb = Workbook()
        ws = wb.active
        ws.column_dimensions['A'].width = 40

        for cat in list(self.categories.keys()):
            ws.cell(row=self.categories[cat], column=1).value = cat

        wb.save(excel_file_name)

    def excel(self, dict):

        '''
        Input info to the excel sheet. Note: First column of excel isn't col=0, it is col=1 and col=1
        is the headers so input starting at col=2
        '''

        col = self.find_empty_col()
        wb = load_workbook(self.filename)
        sheet = wb.active
        for i in range(col, col+len(self.stocks)):
            for cat in list(self.categories.keys()):
                sheet.column_dimensions[get_column_letter(i)].width=10
                sheet.cell(row=self.categories[cat], column=i).value = dict[self.stocks[i-col]][cat]

        wb.save(self.filename)

    def find_empty_col(self):

        #Finds an empty column in the file, used in excel() method

        col = 1
        wb = load_workbook(self.filename)
        sheet = wb.active
        num_cols = len((list(self.categories.keys())))
        while True:
            if(sheet.cell(row=1, column=col).value == None):
                found = False
                for i in range(num_cols):
                    if(sheet.cell(row=i+1, column=col).value != None):
                        found = False
                        break
                    found = True

                if(found == True):
                    return col
            col += 1

        return -1

    def easy_order(self, file=''):

        '''
        This is an easier way of ordering the categories dictionary. Create a file called easy order.xlsx (or if you decide to name it
        differently, pass it as a param), and in column A order the categories however you like. If they are named properly, this function
        will print the dictionary associated with the order, and all you have to do is delete the dictionary of self.categories and paste
        this one instead.
        '''

        row = 1
        cats = list(self.categories.keys())
        if(file == ''):
            file = 'easy order.xlsx'
        wb = self.import_workbook(file)
        #print(wb.sheetnames)
        order_sheet = wb[wb.sheetnames[0]]
        while True:
            if(order_sheet.cell(row=row, column=1).value == None):
                break
            #print("Row: {} Value: {}".format(row, order_sheet.cell(row=row, column=1).value))
            if(order_sheet.cell(row=row, column=1).value not in cats):
                print("The key '{}' is mispelled or not in the original spreadsheet (row {})".format(order_sheet.cell(row=row, column=1).value, row))
                return
            self.categories[order_sheet.cell(row=row, column=1).value] = row
            row += 1

        print(self.categories)



if __name__ =="__main__":
    chrome = StockScraper(['AAPL'], "a")